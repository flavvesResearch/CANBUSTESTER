from __future__ import annotations

import asyncio
import io
import logging
import math
import os
import sys
import threading
import time
import unicodedata
from contextlib import redirect_stdout
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Literal

import openpyxl
from fastapi import File, FastAPI, HTTPException, Query, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field, validator

from .can_manager import CANManager, CANNotConfiguredError
from .dbc_manager import DBCManager, DBCNotLoadedError
from .log_manager import RecordingManager
from .translations import get_all_translations, get_translation

try:
    import can
except ImportError as exc:  # pragma: no cover - handled during runtime
    raise RuntimeError(
        "python-can is required for CAN operations. Please install dependencies from requirements.txt."
    ) from exc

logger = logging.getLogger(__name__)


def _resource_path(*parts: str) -> Path:
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        base = Path(sys._MEIPASS)
    else:
        base = Path(__file__).resolve().parent.parent
    return base.joinpath(*parts)


def _detect_can_interfaces() -> list[Dict[str, Any]]:
    detector = getattr(can, "detect_available_configs", None)
    if detector is None:  # pragma: no cover - depends on python-can version
        logger.warning("python-can detect_available_configs() fonksiyonu kullanılamıyor.")
        return []

    try:
        with redirect_stdout(io.StringIO()):
            configs = detector()
    except Exception as exc:  # pragma: no cover - depends on environment
        logger.exception("CAN arayüzleri listelenirken hata oluştu: %s", exc)
        return []

    results: list[Dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for config in configs:
        interface = config.get("interface")
        channel = config.get("channel")
        if not interface or channel is None:
            continue
        key = (interface, str(channel))
        if key in seen:
            continue
        seen.add(key)
        kwargs = {k: v for k, v in config.items() if k not in {"interface", "channel"} and v is not None}
        results.append(
            {
                "interface": interface,
                "channel": str(channel),
                "kwargs": kwargs,
            }
        )

    results.sort(key=lambda item: (item["interface"], item["channel"]))
    return results


class FaultInjectionManager:
    """Manages fault injection testing for CAN messages."""

    def __init__(
        self,
        dbc_manager: DBCManager,
        can_manager: CANManager,
    ) -> None:
        self._dbc_manager = dbc_manager
        self._can_manager = can_manager
        self._tasks: Dict[str, Dict[str, Any]] = {}
        self._lock = threading.Lock()
        self._notifier: Optional[Callable[[str, str, Dict[str, Any]], None]] = None

    def set_notifier(self, notifier: Callable[[str, str, Dict[str, Any]], None]) -> None:
        self._notifier = notifier

    def start(
        self,
        message_name: str,
        fault_type: str,
        interval_seconds: float,
        count: int,
        **kwargs: Any,
    ) -> Dict[str, Any]:
        if interval_seconds <= 0:
            raise ValueError("Interval must be greater than zero.")
        if count <= 0:
            raise ValueError("Count must be greater than zero.")

        with self._lock:
            if message_name in self._tasks:
                raise RuntimeError("Fault test already running for this message.")

        try:
            message = self._dbc_manager.get_message_by_name(message_name)
        except KeyError as exc:
            raise RuntimeError("Message not found.") from exc

        status = self._can_manager.get_status()
        if not status.get("configured"):
            raise RuntimeError("Configure CAN interface first.")

        stop_event = threading.Event()
        started_at = time.time()

        info: Dict[str, Any] = {
            "messageName": message_name,
            "faultType": fault_type,
            "intervalSeconds": interval_seconds,
            "totalCount": count,
            "sentCount": 0,
            "startedAt": started_at,
        }

        thread = threading.Thread(
            target=self._run_fault_test,
            args=(message_name, message, fault_type, interval_seconds, count, stop_event, kwargs),
            name=f"fault-injection-{message_name}",
            daemon=True,
        )

        with self._lock:
            self._tasks[message_name] = {
                "thread": thread,
                "stop": stop_event,
                "info": info,
            }

        thread.start()
        return info.copy()

    def stop(self, message_name: str) -> Dict[str, Any]:
        with self._lock:
            task = self._tasks.get(message_name)
            if not task:
                raise RuntimeError("No active fault test for this message.")
            stop_event: threading.Event = task["stop"]
            info = task["info"].copy()

        stop_event.set()
        task["thread"].join(timeout=1.0)

        with self._lock:
            self._tasks.pop(message_name, None)

        return info

    def get_status(self, message_name: Optional[str] = None) -> List[Dict[str, Any]]:
        with self._lock:
            if message_name:
                task = self._tasks.get(message_name)
                return [task["info"].copy()] if task else []
            return [task["info"].copy() for task in self._tasks.values()]

    def stop_all(self) -> None:
        for message_name in list(self._tasks.keys()):
            try:
                self.stop(message_name)
            except RuntimeError:
                continue

    def _run_fault_test(
        self,
        message_name: str,
        message: Any,
        fault_type: str,
        interval_seconds: float,
        count: int,
        stop_event: threading.Event,
        kwargs: Dict[str, Any],
    ) -> None:
        import random

        try:
            for i in range(count):
                if stop_event.is_set():
                    break

                try:
                    if fault_type == "bit-flip":
                        encoded = self._inject_bit_flip(message_name, message, kwargs.get("bit_flip_count", 1))
                    elif fault_type == "dlc-mismatch":
                        encoded = self._inject_dlc_mismatch(message_name, message, kwargs.get("dlc_value", 8))
                    elif fault_type == "out-of-range":
                        encoded = self._inject_out_of_range(
                            message_name,
                            message,
                            kwargs.get("target_signal"),
                            kwargs.get("range_multiplier", 2.0),
                        )
                    elif fault_type == "random-data":
                        encoded = self._inject_random_data(message)
                    elif fault_type == "zero-data":
                        encoded = self._inject_zero_data(message)
                    elif fault_type == "max-data":
                        encoded = self._inject_max_data(message)
                    else:
                        raise ValueError(f"Unknown fault type: {fault_type}")

                    self._send(message_name, encoded, fault_type=fault_type)
                    self._update_progress(message_name, i + 1)
                    self._notify_progress(message_name, i + 1, count)

                except Exception as exc:
                    logger.exception("Fault injection error: %s", exc)

                if stop_event.wait(interval_seconds):
                    break
        finally:
            # Clean up task when completed
            with self._lock:
                self._tasks.pop(message_name, None)
            self._notify_completed(message_name)

    def _inject_bit_flip(self, message_name: str, message: Any, bit_count: int) -> Dict[str, Any]:
        import random

        # Get normal encoded message
        signal_values: Dict[str, Any] = {}
        for signal in message.signals:
            signal_values[signal.name] = getattr(signal, "initial", 0) if signal.initial is not None else 0

        encoded = self._dbc_manager.encode(message_name, signal_values)
        data = bytearray(encoded["data"])

        # Flip random bits
        total_bits = len(data) * 8
        bits_to_flip = min(bit_count, total_bits)
        bit_positions = random.sample(range(total_bits), bits_to_flip)

        for bit_pos in bit_positions:
            byte_index = bit_pos // 8
            bit_offset = bit_pos % 8
            data[byte_index] ^= (1 << bit_offset)

        encoded["data"] = bytes(data)
        encoded["faultType"] = "bit-flip"
        encoded["faultInfo"] = f"{bits_to_flip} bit(s) flipped"
        return encoded

    def _inject_dlc_mismatch(self, message_name: str, message: Any, dlc_value: int) -> Dict[str, Any]:
        # Get normal encoded message
        signal_values: Dict[str, Any] = {}
        for signal in message.signals:
            signal_values[signal.name] = getattr(signal, "initial", 0) if signal.initial is not None else 0

        encoded = self._dbc_manager.encode(message_name, signal_values)
        
        # Change DLC to mismatch actual data length
        actual_len = len(encoded["data"])
        encoded["dlc"] = dlc_value
        
        # Pad or truncate data if needed
        if dlc_value > actual_len:
            encoded["data"] = encoded["data"] + bytes([0] * (dlc_value - actual_len))
        elif dlc_value < actual_len:
            encoded["data"] = encoded["data"][:dlc_value]
        
        encoded["faultType"] = "dlc-mismatch"
        encoded["faultInfo"] = f"DLC={dlc_value}, Data length={actual_len}"
        return encoded

    def _inject_out_of_range(
        self,
        message_name: str,
        message: Any,
        target_signal: Optional[str],
        multiplier: float,
    ) -> Dict[str, Any]:
        if not target_signal:
            raise ValueError("Target signal required for out-of-range fault.")

        signal_values: Dict[str, Any] = {}
        fault_info = ""
        
        for signal in message.signals:
            if signal.name == target_signal:
                # Set value beyond max limit
                if signal.maximum is not None:
                    value = signal.maximum * multiplier
                    fault_info = f"{signal.name}={value} (max={signal.maximum})"
                elif signal.minimum is not None:
                    value = signal.minimum * multiplier
                    fault_info = f"{signal.name}={value} (min={signal.minimum})"
                else:
                    value = 999999  # Arbitrary large value
                    fault_info = f"{signal.name}={value} (no limits defined)"
                signal_values[signal.name] = value
            else:
                signal_values[signal.name] = getattr(signal, "initial", 0) if signal.initial is not None else 0

        try:
            encoded = self._dbc_manager.encode(message_name, signal_values)
        except Exception as exc:
            # If encoding fails due to out of range, use raw data approach
            encoded = {
                "arbitration_id": message.frame_id,
                "data": bytes([0xFF] * int(message.length or 8)),
                "is_extended": message.is_extended_frame,
                "dlc": int(message.length or 8),
            }
            fault_info += f" (encoding failed: {exc})"

        encoded["faultType"] = "out-of-range"
        encoded["faultInfo"] = fault_info
        return encoded

    def _inject_random_data(self, message: Any) -> Dict[str, Any]:
        import random
        
        byte_length = int(message.length or 8)
        data = bytes([random.randint(0, 255) for _ in range(byte_length)])
        
        return {
            "arbitration_id": message.frame_id,
            "data": data,
            "is_extended": message.is_extended_frame,
            "dlc": byte_length,
            "faultType": "random-data",
            "faultInfo": "All bytes randomized",
        }

    def _inject_zero_data(self, message: Any) -> Dict[str, Any]:
        byte_length = int(message.length or 8)
        data = bytes([0x00] * byte_length)
        
        return {
            "arbitration_id": message.frame_id,
            "data": data,
            "is_extended": message.is_extended_frame,
            "dlc": byte_length,
            "faultType": "zero-data",
            "faultInfo": "All bytes set to 0x00",
        }

    def _inject_max_data(self, message: Any) -> Dict[str, Any]:
        byte_length = int(message.length or 8)
        data = bytes([0xFF] * byte_length)
        
        return {
            "arbitration_id": message.frame_id,
            "data": data,
            "is_extended": message.is_extended_frame,
            "dlc": byte_length,
            "faultType": "max-data",
            "faultInfo": "All bytes set to 0xFF",
        }

    def _send(self, message_name: str, encoded: Dict[str, Any], fault_type: str) -> None:
        message = _build_can_message(encoded)
        self._can_manager.send(message)
        if self._notifier:
            payload = dict(encoded)
            payload["faultType"] = fault_type
            if "faultInfo" in encoded:
                payload["faultInfo"] = encoded["faultInfo"]
            self._notifier(message_name, "tx", payload)

    def _update_progress(self, message_name: str, sent_count: int) -> None:
        with self._lock:
            task = self._tasks.get(message_name)
            if task:
                task["info"]["sentCount"] = sent_count

    def _notify_progress(self, message_name: str, sent_count: int, total_count: int) -> None:
        if self._notifier:
            self._notifier(
                message_name,
                "fault",
                {
                    "status": "progress",
                    "sentCount": sent_count,
                    "totalCount": total_count,
                    "messageName": message_name,
                },
            )

    def _notify_completed(self, message_name: str) -> None:
        if self._notifier:
            self._notifier(
                message_name,
                "fault",
                {
                    "status": "completed",
                    "messageName": message_name,
                },
            )


class SignalChaserManager:
    """Manages automated signal and error-code emission workflows."""

    MAX_CODES = 4096

    def __init__(
        self,
        dbc_manager: DBCManager,
        can_manager: CANManager,
    ) -> None:
        self._dbc_manager = dbc_manager
        self._can_manager = can_manager
        self._tasks: Dict[str, Dict[str, Any]] = {}
        self._lock = threading.Lock()
        self._notifier: Optional[Callable[[str, Dict[str, Any]], None]] = None

    def set_notifier(self, notifier: Callable[[str, Dict[str, Any]], None]) -> None:
        self._notifier = notifier

    def start(
        self,
        message_name: str,
        interval_seconds: float,
        *,
        mode: str = "signals",
        codes: Optional[List[int]] = None,
        source: Optional[str] = None,
        descriptions: Optional[Dict[int, str]] = None,
        target_signal: Optional[str] = None,
    ) -> Dict[str, Any]:
        if interval_seconds <= 0:
            raise ValueError("Süre 0'dan büyük olmalı.")

        with self._lock:
            if message_name in self._tasks:
                raise RuntimeError("Bu mesaj için sinyal taraması zaten çalışıyor.")

        if mode == "signals":
            return self._start_signal_scan(message_name, interval_seconds)
        if mode == "codes":
            if not codes:
                raise ValueError("Gönderilecek hata kodu bulunamadı.")
            # Check if this is decimal mode (source == "excel-decimal" and target_signal is set)
            if source == "excel-decimal" and target_signal:
                return self._start_code_scan_decimal(
                    message_name,
                    interval_seconds,
                    codes,
                    target_signal,
                    source=source,
                    descriptions=descriptions,
                )
            return self._start_code_scan(
                message_name,
                interval_seconds,
                codes,
                source=source,
                descriptions=descriptions,
            )
        raise ValueError("Geçersiz tarama modu.")

    def _start_signal_scan(self, message_name: str, interval_seconds: float) -> Dict[str, Any]:
        try:
            message = self._dbc_manager.get_message_by_name(message_name)
        except KeyError as exc:
            raise RuntimeError("Mesaj bulunamadı.") from exc
        signals = list(message.signals)
        if not signals:
            raise RuntimeError("Mesajda sinyal bulunamadı.")

        status = self._can_manager.get_status()
        if not status.get("configured"):
            raise RuntimeError("Önce CAN arayüzünü yapılandırın.")

        stop_event = threading.Event()
        started_at = time.time()
        info: Dict[str, Any] = {
            "messageName": message_name,
            "intervalSeconds": interval_seconds,
            "startedAt": started_at,
            "mode": "signals",
            "signals": [signal.name for signal in signals],
            "currentSignal": None,
            "currentCode": None,
            "currentIndex": None,
        }

        thread = threading.Thread(
            target=self._run_signal_scan,
            args=(message_name, signals, interval_seconds, stop_event),
            name=f"signal-chaser-{message_name}",
            daemon=True,
        )

        with self._lock:
            self._tasks[message_name] = {
                "thread": thread,
                "stop": stop_event,
                "info": info,
            }

        thread.start()
        return info.copy()

    def _start_code_scan(
        self,
        message_name: str,
        interval_seconds: float,
        codes: List[int],
        *,
        source: Optional[str] = None,
        descriptions: Optional[Dict[int, str]] = None,
    ) -> Dict[str, Any]:
        try:
            message = self._dbc_manager.get_message_by_name(message_name)
        except KeyError as exc:
            raise RuntimeError("Mesaj bulunamadı.") from exc

        status = self._can_manager.get_status()
        if not status.get("configured"):
            raise RuntimeError("Önce CAN arayüzünü yapılandırın.")

        if len(codes) > self.MAX_CODES:
            raise ValueError(f"En fazla {self.MAX_CODES} hata kodu gönderilebilir.")

        filtered_codes = [code for code in codes if code is not None]
        if not filtered_codes:
            raise ValueError("Geçerli hata kodu bulunamadı.")

        stop_event = threading.Event()
        started_at = time.time()
        byte_length = int(message.length or 8)
        if byte_length <= 0:
            byte_length = 8
        code_lookup: Dict[int, str] = {}
        if descriptions:
            for code, text in descriptions.items():
                if isinstance(code, int) and isinstance(text, str):
                    code_lookup[code] = text.strip()

        info: Dict[str, Any] = {
            "messageName": message_name,
            "intervalSeconds": interval_seconds,
            "startedAt": started_at,
            "mode": "codes",
            "codeSource": source,
            "codeCount": len(filtered_codes),
            "codePreview": [self._format_code(code, byte_length * 2) for code in filtered_codes[:5]],
            "codeHexWidth": byte_length * 2,
            "currentSignal": None,
            "currentCode": None,
            "currentIndex": None,
            "currentDescription": None,
        }
        if code_lookup:
            info["codeDescriptions"] = {
                self._format_code(code, byte_length * 2): text for code, text in code_lookup.items()
            }
            info["codeDescriptionsCount"] = len(code_lookup)

        thread = threading.Thread(
            target=self._run_code_scan,
            args=(
                message_name,
                message,
                tuple(filtered_codes),
                interval_seconds,
                stop_event,
                code_lookup,
            ),
            name=f"code-chaser-{message_name}",
            daemon=True,
        )

        with self._lock:
            self._tasks[message_name] = {
                "thread": thread,
                "stop": stop_event,
                "info": info,
                "descriptions": code_lookup,
            }

        thread.start()
        return info.copy()

    def _start_code_scan_decimal(
        self,
        message_name: str,
        interval_seconds: float,
        codes: List[int],
        target_signal: str,
        *,
        source: Optional[str] = None,
        descriptions: Optional[Dict[int, str]] = None,
    ) -> Dict[str, Any]:
        """Start code scan using decimal values assigned to a specific signal."""
        try:
            message = self._dbc_manager.get_message_by_name(message_name)
        except KeyError as exc:
            raise RuntimeError("Mesaj bulunamadı.") from exc

        # Verify target signal exists
        signal_names = [signal.name for signal in message.signals]
        if target_signal not in signal_names:
            raise RuntimeError(f"Sinyal '{target_signal}' bu mesajda bulunamadı.")

        status = self._can_manager.get_status()
        if not status.get("configured"):
            raise RuntimeError("Önce CAN arayüzünü yapılandırın.")

        if len(codes) > self.MAX_CODES:
            raise ValueError(f"En fazla {self.MAX_CODES} hata kodu gönderilebilir.")

        filtered_codes = [code for code in codes if code is not None]
        if not filtered_codes:
            raise ValueError("Geçerli hata kodu bulunamadı.")

        stop_event = threading.Event()
        started_at = time.time()
        code_lookup: Dict[int, str] = {}
        if descriptions:
            for code, text in descriptions.items():
                if isinstance(code, int) and isinstance(text, str):
                    code_lookup[code] = text.strip()

        info: Dict[str, Any] = {
            "messageName": message_name,
            "intervalSeconds": interval_seconds,
            "startedAt": started_at,
            "mode": "codes",
            "codeSource": source,
            "targetSignal": target_signal,
            "codeCount": len(filtered_codes),
            "codePreview": filtered_codes[:5],
            "currentSignal": target_signal,
            "currentCode": None,
            "currentIndex": None,
            "currentDescription": None,
        }
        if code_lookup:
            info["codeDescriptions"] = code_lookup
            info["codeDescriptionsCount"] = len(code_lookup)

        thread = threading.Thread(
            target=self._run_code_scan_decimal,
            args=(
                message_name,
                message,
                tuple(filtered_codes),
                target_signal,
                interval_seconds,
                stop_event,
                code_lookup,
            ),
            name=f"code-decimal-chaser-{message_name}",
            daemon=True,
        )

        with self._lock:
            self._tasks[message_name] = {
                "thread": thread,
                "stop": stop_event,
                "info": info,
                "descriptions": code_lookup,
            }

        thread.start()
        return info.copy()

    def stop(self, message_name: str) -> Dict[str, Any]:
        with self._lock:
            task = self._tasks.get(message_name)
            if not task:
                raise RuntimeError("Bu mesaj için aktif sinyal taraması yok.")
            stop_event: threading.Event = task["stop"]
            info = task["info"].copy()

        stop_event.set()
        task["thread"].join(timeout=1.0)

        with self._lock:
            self._tasks.pop(message_name, None)

        return info

    def get_status(self) -> List[Dict[str, Any]]:
        with self._lock:
            return [task["info"].copy() for task in self._tasks.values()]

    def stop_all(self) -> None:
        for message_name in list(self._tasks.keys()):
            try:
                self.stop(message_name)
            except RuntimeError:
                continue

    def _run_signal_scan(
        self,
        message_name: str,
        signals: List[Any],
        interval_seconds: float,
        stop_event: threading.Event,
    ) -> None:
        index = 0
        signal_count = len(signals)

        while not stop_event.is_set():
            payload: Dict[str, Any] = {}
            for signal in signals:
                payload[signal.name] = self._min_value(signal)

            current_signal = signals[index]
            payload[current_signal.name] = self._max_value(current_signal)

            try:
                encoded = self._dbc_manager.encode(message_name, payload)
                self._send(message_name, encoded, mode="signals")
                self._update_current_signal(message_name, current_signal.name)
            except Exception as exc:  # pragma: no cover - runtime safety
                logger.exception("Sinyal taraması gönderim hatası: %s", exc)

            if stop_event.wait(interval_seconds):
                break
            index = (index + 1) % signal_count

        self._update_current_signal(message_name, None)

    def _run_code_scan(
        self,
        message_name: str,
        message: Any,
        codes: tuple[int, ...],
        interval_seconds: float,
        stop_event: threading.Event,
        descriptions: Dict[int, str],
    ) -> None:
        index = 0
        code_count = len(codes)
        byte_length = int(message.length or 8)
        if byte_length <= 0:
            byte_length = 8

        while not stop_event.is_set():
            code = codes[index]
            description = descriptions.get(code) if descriptions else None

            try:
                encoded = self._encode_code_payload(message, code, byte_length)
                encoded["code"] = code
                if description:
                    encoded["description"] = description
                self._send(message_name, encoded, mode="codes")
                self._update_current_code(message_name, code, index, byte_length * 2, description)
            except Exception as exc:  # pragma: no cover - runtime safety
                logger.exception("Hata kodu taraması gönderim hatası: %s", exc)

            if stop_event.wait(interval_seconds):
                break
            index = (index + 1) % code_count

        self._update_current_code(message_name, None, None, byte_length * 2, None)

    def _run_code_scan_decimal(
        self,
        message_name: str,
        message: Any,
        codes: tuple[int, ...],
        target_signal: str,
        interval_seconds: float,
        stop_event: threading.Event,
        descriptions: Dict[int, str],
    ) -> None:
        """Run code scan by assigning decimal values to a specific signal."""
        index = 0
        code_count = len(codes)

        while not stop_event.is_set():
            code = codes[index]
            description = descriptions.get(code) if descriptions else None

            try:
                # Build signal payload: set target signal to decimal code value
                # All other signals set to their minimum or initial values
                signal_values: Dict[str, Any] = {}
                for signal in message.signals:
                    if signal.name == target_signal:
                        signal_values[signal.name] = code
                    else:
                        signal_values[signal.name] = self._min_value(signal)

                # Encode using DBC
                encoded = self._dbc_manager.encode(message_name, signal_values)
                encoded["code"] = code
                if description:
                    encoded["description"] = description
                self._send(message_name, encoded, mode="codes")
                # Use hex width of 1 for decimal display (just show as decimal number)
                self._update_current_code(message_name, code, index, 1, description)
            except Exception as exc:  # pragma: no cover - runtime safety
                logger.exception("Hata kodu decimal taraması gönderim hatası: %s", exc)

            if stop_event.wait(interval_seconds):
                break
            index = (index + 1) % code_count

        self._update_current_code(message_name, None, None, 1, None)

    def _encode_code_payload(self, message: Any, code: int, byte_length: int) -> Dict[str, Any]:
        if code < 0:
            raise ValueError("Hata kodu negatif olamaz.")

        max_value = (1 << (byte_length * 8)) - 1
        if code > max_value:
            raise ValueError("Hata kodu mesaj uzunluğunu aşıyor.")

        data = code.to_bytes(byte_length, byteorder="big")
        return {
            "arbitration_id": message.frame_id,
            "data": data,
            "is_extended": message.is_extended_frame,
            "dlc": byte_length,
        }

    def _send(self, message_name: str, encoded: Dict[str, Any], *, mode: str) -> None:
        message = _build_can_message(encoded)
        self._can_manager.send(message)
        if self._notifier:
            payload = dict(encoded)
            payload.setdefault("mode", mode)
            if "description" in encoded:
                payload["description"] = encoded["description"]
            self._notifier(message_name, payload)

    def _update_current_signal(self, message_name: str, signal_name: Optional[str]) -> None:
        with self._lock:
            task = self._tasks.get(message_name)
            if not task:
                return
            task["info"]["currentSignal"] = signal_name
            if signal_name is not None:
                task["info"]["lastSentAt"] = time.time()
            else:
                task["info"]["lastSentAt"] = time.time()

    def _update_current_code(
        self,
        message_name: str,
        code_value: Optional[int],
        index: Optional[int],
        hex_width: int,
        description: Optional[str],
    ) -> None:
        with self._lock:
            task = self._tasks.get(message_name)
            if not task:
                return
            info = task["info"]
            info["currentIndex"] = index
            if code_value is None:
                info["currentCode"] = None
            else:
                info["currentCode"] = self._format_code(code_value, hex_width)
                info["lastSentAt"] = time.time()
            info["currentDescription"] = description
            if code_value is None:
                info["lastSentAt"] = time.time()

    @staticmethod
    def _min_value(signal: Any) -> float:
        if signal.minimum is not None:
            return signal.minimum
        if signal.initial is not None:
            return signal.initial
        if signal.choices:
            try:
                keys = list(signal.choices.keys())
                keys.sort(key=lambda item: float(getattr(item, "value", item)))
                best = keys[0]
                return getattr(best, "value", best)
            except Exception:
                return 0.0
        return 0.0

    @staticmethod
    def _max_value(signal: Any) -> float:
        if signal.maximum is not None:
            return signal.maximum
        if signal.choices:
            try:
                keys = list(signal.choices.keys())
                keys.sort(key=lambda item: float(getattr(item, "value", item)))
                best = keys[-1]
                return getattr(best, "value", best)
            except Exception:
                return 1.0
        if signal.initial is not None:
            return signal.initial
        length = getattr(signal, "length", None)
        if isinstance(length, int) and length > 0:
            return float((1 << length) - 1)
        return 1.0

    @staticmethod
    def _format_code(code: int, hex_width: int) -> str:
        # If hex_width is 1, format as decimal (for decimal mode)
        if hex_width == 1:
            return str(code)
        # Otherwise format as hex
        width = max(2, hex_width)
        return f"0x{code:0{width}X}"
class MessageBroadcaster:
    """Tracks websocket connections and sends events to all clients."""

    def __init__(self, buffer_delay_seconds: float = 0.1) -> None:
        self._connections: set[WebSocket] = set()
        self._lock = asyncio.Lock()
        self._loop: Optional[asyncio.AbstractEventLoop] = None
        self._queue: Optional[asyncio.Queue[Dict[str, Any]]] = None
        self._broadcast_task: Optional[asyncio.Task[None]] = None
        self._buffer_delay_seconds = buffer_delay_seconds

    def set_loop(self, loop: asyncio.AbstractEventLoop) -> None:
        self._loop = loop
        self._queue = asyncio.Queue()
        self._broadcast_task = loop.create_task(self._broadcast_loop())

    async def shutdown(self) -> None:
        if self._broadcast_task:
            self._broadcast_task.cancel()
            try:
                await self._broadcast_task
            except asyncio.CancelledError:
                pass
        # Send any remaining messages
        if self._queue and not self._queue.empty():
            batch = []
            while not self._queue.empty():
                batch.append(self._queue.get_nowait())
            if batch:
                await self.broadcast(batch)

    async def _broadcast_loop(self) -> None:
        while True:
            try:
                await asyncio.sleep(self._buffer_delay_seconds)
                if self._queue and not self._queue.empty():
                    batch = []
                    while not self._queue.empty():
                        batch.append(self._queue.get_nowait())
                    if batch:
                        await self.broadcast(batch)
            except asyncio.CancelledError:
                break
            except Exception:
                logger.exception("Error in broadcast loop")

    async def connect(self, websocket: WebSocket) -> None:
        await websocket.accept()
        async with self._lock:
            self._connections.add(websocket)

    async def disconnect(self, websocket: WebSocket) -> None:
        async with self._lock:
            self._connections.discard(websocket)

    async def broadcast(self, batch: List[Dict[str, Any]]) -> None:
        if not batch:
            return
        
        payload: Dict[str, Any]
        if len(batch) == 1:
            payload = batch[0]
        else:
            payload = {"type": "batch", "messages": batch}

        async with self._lock:
            dead: list[WebSocket] = []
            for connection in self._connections:
                try:
                    await connection.send_json(payload)
                except Exception:
                    dead.append(connection)
            for connection in dead:
                self._connections.discard(connection)

    def send_threadsafe(self, payload: Dict[str, Any]) -> None:
        if self._queue is None:
            logger.debug("No event loop set for broadcaster, dropping payload")
            return
        try:
            self._queue.put_nowait(payload)
        except asyncio.QueueFull:
            logger.warning("Broadcaster queue is full, dropping payload.")


app = FastAPI(title="CAN Bus Tester", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
STATIC_DIR = _resource_path("static")
INDEX_FILE = STATIC_DIR / "index.html"
PLAYBACK_FILE = STATIC_DIR / "playback.html"

app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

dbc_manager = DBCManager()
can_manager = CANManager()
broadcaster = MessageBroadcaster()

env_record_dir = os.environ.get("CANBUS_RECORDINGS_DIR")
if env_record_dir:
    recording_root = Path(env_record_dir).expanduser()
else:
    recording_root = Path.cwd() / "recordings"
recording_root.mkdir(parents=True, exist_ok=True)

recording_manager = RecordingManager(recording_root)
signal_chaser = SignalChaserManager(dbc_manager, can_manager)
fault_injection = FaultInjectionManager(dbc_manager, can_manager)


class ConfigureInterfaceRequest(BaseModel):
    interface: str
    channel: str
    bitrate: Optional[int] = Field(default=None, ge=0)
    kwargs: Dict[str, Any] = Field(default_factory=dict)


class MessageSendRequest(BaseModel):
    message_name: str = Field(alias="messageName")
    signals: Dict[str, Any]
    period_ms: Optional[int] = Field(default=None, alias="periodMs", gt=0)
    task_key: Optional[str] = Field(default=None, alias="taskKey")

    @validator("task_key", always=True)
    def default_task_key(cls, value: Optional[str], values: Dict[str, Any]) -> Optional[str]:
        period = values.get("period_ms")
        if period and not value:
            return values.get("message_name")
        return value


class StopTaskRequest(BaseModel):
    task_key: str = Field(alias="taskKey")


class RecordingStartRequest(BaseModel):
    name: Optional[str] = None


class SignalChaserStartRequest(BaseModel):
    message_name: str = Field(alias="messageName")
    interval_seconds: float = Field(alias="intervalSeconds", gt=0)
    mode: Literal["signals", "codes"] = Field(default="signals")
    code_source: Optional[Literal["excel", "excel-decimal", "manual"]] = Field(default=None, alias="codeSource")
    codes: Optional[List[Any]] = Field(default=None)
    target_signal: Optional[str] = Field(default=None, alias="targetSignal")
    code_range_start: Optional[str] = Field(default=None, alias="codeRangeStart")
    code_range_end: Optional[str] = Field(default=None, alias="codeRangeEnd")
    code_descriptions: Optional[Dict[str, Any]] = Field(default=None, alias="codeDescriptions")


class SignalChaserStopRequest(BaseModel):
    message_name: str = Field(alias="messageName")


class FaultInjectionStartRequest(BaseModel):
    message_name: str = Field(alias="messageName")
    fault_type: Literal["bit-flip", "dlc-mismatch", "out-of-range", "random-data", "zero-data", "max-data"] = Field(alias="faultType")
    interval_seconds: float = Field(alias="intervalSeconds", gt=0)
    count: int = Field(gt=0)
    bit_flip_count: Optional[int] = Field(default=1, alias="bitFlipCount", ge=1, le=64)
    dlc_value: Optional[int] = Field(default=None, alias="dlcValue", ge=0, le=8)
    target_signal: Optional[str] = Field(default=None, alias="targetSignal")
    range_multiplier: Optional[float] = Field(default=2.0, alias="rangeMultiplier", ge=1.1)


class FaultInjectionStopRequest(BaseModel):
    message_name: str = Field(alias="messageName")


def _build_can_message(encoded: Dict[str, Any]) -> can.Message:
    return can.Message(
        arbitration_id=encoded["arbitration_id"],
        data=encoded["data"],
        dlc=encoded["dlc"],
        is_extended_id=encoded["is_extended"],
    )


def _parse_code_value(raw: Any) -> int:
    if raw is None:
        raise ValueError("Boş değer.")
    if isinstance(raw, bool):
        raise ValueError("Geçersiz değer.")
    if isinstance(raw, int):
        if raw < 0:
            raise ValueError("Hata kodu negatif olamaz.")
        return raw
    if isinstance(raw, float):
        if math.isnan(raw) or math.isinf(raw) or not raw.is_integer():
            raise ValueError("Hata kodu sayısı geçersiz.")
        value = int(raw)
        if value < 0:
            raise ValueError("Hata kodu negatif olamaz.")
        return value

    text = str(raw).strip()
    if not text:
        raise ValueError("Boş değer.")
    cleaned = text.replace(" ", "")
    if cleaned.lower().startswith("0x"):
        cleaned = cleaned[2:]
    if cleaned.endswith(("h", "H")):
        cleaned = cleaned[:-1]
    cleaned = cleaned.replace("_", "")

    try:
        value = int(cleaned, 16)
    except ValueError as exc:
        try:
            value = int(cleaned, 10)
        except ValueError as second_exc:
            raise ValueError(f"'{text}' değeri hata koduna dönüştürülemedi.") from second_exc

    if value < 0:
        raise ValueError("Hata kodu negatif olamaz.")
    return value


def _handle_tx_event(
    message_name: str,
    encoded: Dict[str, Any],
    task_key: Optional[str] = None,
    period_ms: Optional[int] = None,
) -> None:
    timestamp = time.time()
    payload = {
        "type": "tx",
        "message": message_name,
        "taskKey": task_key,
        "periodMs": period_ms,
        "id": encoded["arbitration_id"],
        "dlc": encoded["dlc"],
        "data": list(encoded["data"]),
        "timestamp": timestamp,
    }
    mode = encoded.get("mode")
    if mode:
        payload["mode"] = mode
    code_value = encoded.get("code")
    if code_value is not None:
        payload["code"] = code_value
    description = encoded.get("description")
    if description:
        payload["description"] = description
    broadcaster.send_threadsafe(payload)
    recording_manager.append_event(payload)


signal_chaser.set_notifier(_handle_tx_event)


def _handle_fault_event(message_name: str, event_type: str, data: Dict[str, Any]) -> None:
    if event_type == "tx":
        timestamp = time.time()
        payload = {
            "type": "tx",
            "message": message_name,
            "id": data["arbitration_id"],
            "dlc": data["dlc"],
            "data": list(data["data"]),
            "timestamp": timestamp,
            "faultType": data.get("faultType"),
            "faultInfo": data.get("faultInfo"),
        }
        broadcaster.send_threadsafe(payload)
        recording_manager.append_event(payload)
    elif event_type == "fault":
        broadcaster.send_threadsafe({"type": "fault", **data})


fault_injection.set_notifier(_handle_fault_event)


def _on_can_message(message: can.Message) -> None:
    payload: Dict[str, Any] = {
        "type": "rx",
        "id": message.arbitration_id,
        "timestamp": message.timestamp if message.timestamp else time.time(),
        "is_extended": message.is_extended_id,
        "dlc": message.dlc,
        "data": list(message.data),
    }
    if dbc_manager.is_loaded():
        try:
            decoded = dbc_manager.decode(message.arbitration_id, message.data)
        except DBCNotLoadedError:
            decoded = None
        if decoded:
            payload["decoded"] = decoded
    broadcaster.send_threadsafe(payload)
    recording_manager.append_event(payload)


can_manager.register_callback(_on_can_message)


@app.on_event("startup")
async def on_startup() -> None:
    broadcaster.set_loop(asyncio.get_running_loop())
    logger.info("CAN Bus Tester API ready")


@app.on_event("shutdown")
async def on_shutdown() -> None:
    await broadcaster.shutdown()
    can_manager.shutdown()
    signal_chaser.stop_all()
    fault_injection.stop_all()


@app.get("/")
async def index() -> FileResponse:
    return FileResponse(str(INDEX_FILE))


@app.get("/playback")
async def playback() -> FileResponse:
    return FileResponse(str(PLAYBACK_FILE))


@app.get("/api/translations/{lang}")
async def get_translations(lang: str) -> Dict[str, Any]:
    """Get all translations for a specific language."""
    translations = get_all_translations(lang)
    return {"language": lang, "translations": translations}


@app.get("/api/interface/available")
async def get_available_interfaces() -> Dict[str, Any]:
    interfaces = _detect_can_interfaces()
    return {"interfaces": interfaces}


@app.get("/api/interface/status")
async def get_interface_status() -> Dict[str, Any]:
    return can_manager.get_status()


@app.post("/api/interface/configure")
async def configure_interface(request: ConfigureInterfaceRequest) -> Dict[str, Any]:
    try:
        status = can_manager.configure(
            interface=request.interface,
            channel=request.channel,
            bitrate=request.bitrate,
            **request.kwargs,
        )
        broadcaster.send_threadsafe({"type": "interface", "status": status})
        return status
    except Exception as exc:
        logger.exception("Failed to configure CAN interface: %s", exc)
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/dbc/load")
async def load_dbc(file: UploadFile = File(...)) -> Dict[str, Any]:
    try:
        contents = await file.read()
        metadata = dbc_manager.load_from_content(contents, file.filename)
        broadcaster.send_threadsafe(
            {"type": "dbc", "path": metadata.get("path"), "name": metadata.get("name")}
        )
        signal_chaser.stop_all()
        return metadata
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:  # pragma: no cover - depends on cantools
        logger.exception("Failed to load DBC: %s", exc)
        raise HTTPException(status_code=400, detail="DBC dosyası yüklenemedi.")


@app.get("/api/dbc/messages")
async def get_dbc_messages() -> Dict[str, Any]:
    if not dbc_manager.is_loaded():
        raise HTTPException(status_code=404, detail="DBC not loaded.")
    return dbc_manager.list_messages()


@app.get("/api/messages/chaser/status")
async def get_signal_chaser_status(message_name: Optional[str] = Query(default=None, alias="messageName")) -> Dict[str, Any]:
    tasks = signal_chaser.get_status()
    if message_name:
        tasks = [task for task in tasks if task.get("messageName") == message_name]
    return {"tasks": tasks}


@app.post("/api/messages/chaser/codes/upload")
async def upload_chaser_codes(file: UploadFile = File(...)) -> Dict[str, Any]:
    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Excel dosyası boş olamaz.")

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as exc:
        logger.exception("Excel dosyası okunamadı: %s", exc)
        raise HTTPException(status_code=400, detail="Excel dosyası okunamadı.")

    sheet = workbook.active
    header_aliases = {
        "hata kodları (hex)",
        "hata kodları hex",
        "hata kodları",
        "hatakodları(hex)",
        "hatakodlari(hex)",
        "error codes (hex)",
        "error codes hex",
        "error codes",
        "errorcodes(hex)",
        "errorcodes",
    }

    def _normalise_header(value: str) -> str:
        compact = " ".join(value.strip().split())
        normalized = unicodedata.normalize("NFKD", compact).lower()
        return "".join(ch for ch in normalized if ch.isalnum())

    normalised_aliases = {_normalise_header(alias) for alias in header_aliases}

    target_column: Optional[int] = None
    header_row_index: Optional[int] = None
    description_column: Optional[int] = None
    description_headers = {
        "hata başlıkları",
        "hatabaşlıkları",
        "hata açıklaması",
        "hata aciklamasi",
        "description",
        "error description",
        "error titles",
    }
    normalised_description_headers = {_normalise_header(alias) for alias in description_headers}

    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), start=1):
        for column_index, cell in enumerate(row, start=1):
            if isinstance(cell, str):
                candidate = _normalise_header(cell)
                if target_column is None and candidate in normalised_aliases:
                    target_column = column_index
                    header_row_index = row_index
                if description_column is None and candidate in normalised_description_headers:
                    description_column = column_index
        if target_column is not None:
            break

    if target_column is None or header_row_index is None:
        # Fallback: try first non-empty column if there is only one column with data
        non_empty_columns = set()
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
            for column_index, cell in enumerate(row, start=1):
                if cell not in (None, "", " "):
                    non_empty_columns.add(column_index)
        if len(non_empty_columns) == 1:
            target_column = non_empty_columns.pop()
            header_row_index = 0
        else:
            raise HTTPException(status_code=400, detail="Excel dosyasında 'HATA KODLARI (hex)' başlığı bulunamadı.")

    codes: List[int] = []
    code_descriptions: Dict[int, str] = {}
    invalid_count = 0

    for row_index in range(header_row_index + 1, sheet.max_row + 1):
        value = sheet.cell(row=row_index, column=target_column).value
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        try:
            code_value = _parse_code_value(value)
        except ValueError:
            invalid_count += 1
            continue
        codes.append(code_value)
        if description_column is not None and row_index <= sheet.max_row:
            description_raw = sheet.cell(row=row_index, column=description_column).value
            if isinstance(description_raw, str) and description_raw.strip():
                code_descriptions[code_value] = description_raw.strip()

    if not codes:
        raise HTTPException(status_code=400, detail="Excel dosyasında geçerli hata kodu bulunamadı.")

    truncated = False
    original_count = len(codes)
    if original_count > SignalChaserManager.MAX_CODES:
        codes = codes[: SignalChaserManager.MAX_CODES]
        truncated = True

    max_bits = max(code.bit_length() for code in codes) if codes else 1
    hex_digits = max(2, (max_bits + 3) // 4)
    if hex_digits % 2 != 0:
        hex_digits += 1
    codes_payload = [f"0x{code:0{hex_digits}X}" for code in codes]

    return {
        "fileName": file.filename,
        "count": len(codes_payload),
        "originalCount": original_count,
        "invalidCount": invalid_count,
        "truncated": truncated,
        "maxAllowed": SignalChaserManager.MAX_CODES,
        "hexDigits": hex_digits,
        "codes": codes_payload,
        "descriptions": {f"0x{code:0{hex_digits}X}": text for code, text in code_descriptions.items()},
    }


@app.post("/api/messages/chaser/codes/upload-decimal")
async def upload_chaser_codes_decimal(file: UploadFile = File(...)) -> Dict[str, Any]:
    """Upload Excel file with hex codes, convert to decimal values for signal assignment."""
    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Excel dosyası boş olamaz.")

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as exc:
        logger.exception("Excel dosyası okunamadı: %s", exc)
        raise HTTPException(status_code=400, detail="Excel dosyası okunamadı.")

    sheet = workbook.active
    header_aliases = {
        "hata kodları (hex)",
        "hata kodları hex",
        "hata kodları",
        "hatakodları(hex)",
        "hatakodlari(hex)",
        "error codes (hex)",
        "error codes hex",
        "error codes",
        "errorcodes(hex)",
        "errorcodes",
    }

    def _normalise_header(value: str) -> str:
        compact = " ".join(value.strip().split())
        normalized = unicodedata.normalize("NFKD", compact).lower()
        return "".join(ch for ch in normalized if ch.isalnum())

    normalised_aliases = {_normalise_header(alias) for alias in header_aliases}

    target_column: Optional[int] = None
    header_row_index: Optional[int] = None
    description_column: Optional[int] = None
    description_headers = {
        "hata başlıkları",
        "hatabaşlıkları",
        "hata açıklaması",
        "hata aciklamasi",
        "description",
        "error description",
        "error titles",
    }
    normalised_description_headers = {_normalise_header(alias) for alias in description_headers}

    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), start=1):
        for column_index, cell in enumerate(row, start=1):
            if isinstance(cell, str):
                candidate = _normalise_header(cell)
                if target_column is None and candidate in normalised_aliases:
                    target_column = column_index
                    header_row_index = row_index
                if description_column is None and candidate in normalised_description_headers:
                    description_column = column_index
        if target_column is not None:
            break

    if target_column is None or header_row_index is None:
        # Fallback: try first non-empty column if there is only one column with data
        non_empty_columns = set()
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
            for column_index, cell in enumerate(row, start=1):
                if cell not in (None, "", " "):
                    non_empty_columns.add(column_index)
        if len(non_empty_columns) == 1:
            target_column = non_empty_columns.pop()
            header_row_index = 0
        else:
            raise HTTPException(status_code=400, detail="Excel dosyasında 'HATA KODLARI (hex)' başlığı bulunamadı.")

    codes: List[int] = []
    code_descriptions: Dict[int, str] = {}
    invalid_count = 0

    for row_index in range(header_row_index + 1, sheet.max_row + 1):
        value = sheet.cell(row=row_index, column=target_column).value
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        try:
            code_value = _parse_code_value(value)
        except ValueError:
            invalid_count += 1
            continue
        codes.append(code_value)
        if description_column is not None and row_index <= sheet.max_row:
            description_raw = sheet.cell(row=row_index, column=description_column).value
            if isinstance(description_raw, str) and description_raw.strip():
                code_descriptions[code_value] = description_raw.strip()

    if not codes:
        raise HTTPException(status_code=400, detail="Excel dosyasında geçerli hata kodu bulunamadı.")

    truncated = False
    original_count = len(codes)
    if original_count > SignalChaserManager.MAX_CODES:
        codes = codes[: SignalChaserManager.MAX_CODES]
        truncated = True

    # Return decimal values as integers (not hex strings)
    codes_payload = codes

    return {
        "fileName": file.filename,
        "count": len(codes_payload),
        "originalCount": original_count,
        "invalidCount": invalid_count,
        "truncated": truncated,
        "maxAllowed": SignalChaserManager.MAX_CODES,
        "codes": codes_payload,
        "descriptions": {code: text for code, text in code_descriptions.items()},
    }


@app.post("/api/messages/chaser/start")
async def start_signal_chaser(request: SignalChaserStartRequest) -> Dict[str, Any]:
    if not dbc_manager.is_loaded():
        raise HTTPException(status_code=400, detail="Önce bir DBC dosyası yükleyin.")
    try:
        if request.mode == "codes":
            codes: List[int] = []
            description_map: Dict[int, str] = {}

            if request.codes:
                for raw in request.codes:
                    try:
                        codes.append(_parse_code_value(raw))
                    except ValueError as exc:
                        raise HTTPException(status_code=400, detail=str(exc)) from exc
            elif request.code_range_start is not None and request.code_range_end is not None:
                try:
                    range_start = _parse_code_value(request.code_range_start)
                    range_end = _parse_code_value(request.code_range_end)
                except ValueError as exc:
                    raise HTTPException(status_code=400, detail=str(exc)) from exc
                if range_end < range_start:
                    range_start, range_end = range_end, range_start
                range_size = (range_end - range_start) + 1
                if range_size > SignalChaserManager.MAX_CODES:
                    raise HTTPException(
                        status_code=400,
                        detail=f"En fazla {SignalChaserManager.MAX_CODES} kod gönderilebilir. Lütfen aralığı daraltın.",
                    )
                codes = [range_start + offset for offset in range(range_size)]
            else:
                raise HTTPException(status_code=400, detail="Excel ya da manuel kod seçeneği belirlenmedi.")

            if not codes:
                raise HTTPException(status_code=400, detail="Gönderilecek hata kodu bulunamadı.")

            if request.code_descriptions:
                for key, value in request.code_descriptions.items():
                    if value in (None, ""):
                        continue
                    try:
                        code_key = _parse_code_value(key)
                    except ValueError:
                        continue
                    if code_key in codes:
                        description_map[code_key] = str(value).strip()

            task = signal_chaser.start(
                request.message_name,
                request.interval_seconds,
                mode="codes",
                codes=codes,
                source=request.code_source,
                descriptions=description_map or None,
                target_signal=request.target_signal,
            )
        else:
            task = signal_chaser.start(request.message_name, request.interval_seconds)
        return {"status": "running", "task": task}
    except (RuntimeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/messages/chaser/stop")
async def stop_signal_chaser(request: SignalChaserStopRequest) -> Dict[str, Any]:
    try:
        task = signal_chaser.stop(request.message_name)
        return {"status": "stopped", "task": task}
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/messages/fault/start")
async def start_fault_injection(request: FaultInjectionStartRequest) -> Dict[str, Any]:
    if not dbc_manager.is_loaded():
        raise HTTPException(status_code=400, detail="Load a DBC file first.")
    
    try:
        kwargs = {}
        if request.fault_type == "bit-flip":
            kwargs["bit_flip_count"] = request.bit_flip_count or 1
        elif request.fault_type == "dlc-mismatch":
            kwargs["dlc_value"] = request.dlc_value if request.dlc_value is not None else 8
        elif request.fault_type == "out-of-range":
            if not request.target_signal:
                raise HTTPException(status_code=400, detail="Target signal required for out-of-range fault.")
            kwargs["target_signal"] = request.target_signal
            kwargs["range_multiplier"] = request.range_multiplier or 2.0
        
        task = fault_injection.start(
            request.message_name,
            request.fault_type,
            request.interval_seconds,
            request.count,
            **kwargs,
        )
        return {"status": "running", "task": task}
    except (RuntimeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/messages/fault/stop")
async def stop_fault_injection(request: FaultInjectionStopRequest) -> Dict[str, Any]:
    try:
        task = fault_injection.stop(request.message_name)
        return {"status": "stopped", "task": task}
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.get("/api/messages/fault/status")
async def get_fault_injection_status(message_name: Optional[str] = Query(default=None, alias="messageName")) -> Dict[str, Any]:
    tasks = fault_injection.get_status(message_name)
    return {"tasks": tasks}


@app.get("/api/logs")
async def list_logs() -> Dict[str, Any]:
    logs = recording_manager.list_recordings()
    active = recording_manager.get_active()
    if active:
        active["duration"] = max(0.0, time.time() - active["started_at"])
    return {"active": active, "logs": logs}


@app.post("/api/logs/start")
async def start_log(request: RecordingStartRequest) -> Dict[str, Any]:
    try:
        info = recording_manager.start(request.name)
        broadcaster.send_threadsafe({"type": "recording", "state": "started", "record": info})
        return info
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.post("/api/logs/stop")
async def stop_log() -> Dict[str, Any]:
    try:
        info = recording_manager.stop()
        broadcaster.send_threadsafe({"type": "recording", "state": "stopped", "record": info})
        return info
    except RuntimeError as exc:
        raise HTTPException(status_code=400, detail=str(exc))


@app.get("/api/logs/{log_id}")
async def get_log(log_id: str) -> Dict[str, Any]:
    data = recording_manager.get_recording(log_id)
    if data is None:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı.")
    return data


@app.post("/api/logs/{log_id}/decode")
async def decode_log(log_id: str, file: UploadFile = File(...)) -> Dict[str, Any]:
    data = recording_manager.get_recording(log_id)
    if data is None:
        raise HTTPException(status_code=404, detail="Kayıt bulunamadı.")

    contents = await file.read()
    temp_manager = DBCManager()
    try:
        temp_manager.load_from_content(contents, file.filename)
    except Exception as exc:  # pragma: no cover - depends on cantools
        raise HTTPException(status_code=400, detail=f"DBC yüklenemedi: {exc}")

    try:
        metadata = temp_manager.list_messages()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"DBC mesajları okunamadı: {exc}")

    signal_catalog: Dict[tuple[str, str], Dict[str, Any]] = {}
    for message in metadata.get("messages", []):
        for signal in message.get("signals", []):
            signal_catalog[(message["name"], signal["name"])] = signal

    started_at = data.get("started_at", 0.0)
    events = data.get("events", [])
    decoded_events: list[Dict[str, Any]] = []
    events_total = len(events)
    max_events = 2000
    max_points = 5000
    series_map: Dict[str, Dict[str, Any]] = {}

    for event in events:
        event_timestamp = float(event.get("timestamp", started_at))
        relative_time = event_timestamp - started_at
        decoded_payload: Optional[Dict[str, Any]] = None
        data_bytes = bytes(event.get("data", []))
        try:
            decoded_payload = temp_manager.decode(event.get("id", 0), data_bytes)
        except Exception:
            decoded_payload = None

        if decoded_payload is None and event.get("message"):
            try:
                message_obj = temp_manager.get_message_by_name(event["message"])
                decoded_signals = message_obj.decode(data_bytes)
                decoded_payload = {
                    "name": message_obj.name,
                    "signals": decoded_signals,
                    "is_extended": message_obj.is_extended_frame,
                    "comment": message_obj.comment,
                }
            except Exception:
                decoded_payload = None

        if decoded_payload and decoded_payload.get("name"):
            message_name = decoded_payload["name"]
            signals = decoded_payload.get("signals", {})
            for signal_name, value in signals.items():
                key = f"{message_name}.{signal_name}"
                entry = series_map.setdefault(
                    key,
                    {
                        "key": key,
                        "message": message_name,
                        "signal": signal_name,
                        "unit": signal_catalog.get((message_name, signal_name), {}).get("unit"),
                        "points": [],
                    },
                )
                numeric_value: Optional[float]
                if isinstance(value, bool):
                    numeric_value = float(int(value))
                elif isinstance(value, (int, float)):
                    numeric_value = float(value)
                else:
                    numeric_value = None
                if numeric_value is not None:
                    entry["points"].append({"timestamp": event_timestamp, "relative": relative_time, "value": numeric_value})

        if len(decoded_events) < max_events:
            decoded_events.append(
                {
                    "type": event.get("type"),
                    "timestamp": event_timestamp,
                    "relative_time": relative_time,
                    "id": event.get("id"),
                    "dlc": event.get("dlc"),
                    "data": event.get("data"),
                    "message": event.get("message"),
                    "decoded": decoded_payload,
                    "periodMs": event.get("periodMs"),
                }
            )

    series = sorted(series_map.values(), key=lambda item: item["key"])
    for entry in series:
        points = entry.get("points", [])
        original_count = len(points)
        if original_count > max_points:
            step = max(1, math.ceil(original_count / max_points))
            entry["points"] = points[::step]
            entry["downsampled"] = True
            entry["original_points"] = original_count
        else:
            entry["downsampled"] = False
            entry["original_points"] = original_count
    duration = 0.0
    if events:
        last_timestamp = max(float(e.get("timestamp", started_at)) for e in events)
        duration = max(0.0, last_timestamp - started_at)

    return {
        "log": {
            "id": data.get("id"),
            "name": data.get("name"),
            "started_at": started_at,
            "ended_at": data.get("ended_at"),
            "duration": duration,
            "event_count": len(events),
        },
        "events": decoded_events,
        "events_shown": len(decoded_events),
        "events_total": events_total,
        "series": series,
    }


@app.post("/api/messages/send")
async def send_message(request: MessageSendRequest) -> JSONResponse:
    if not dbc_manager.is_loaded():
        raise HTTPException(status_code=400, detail="DBC file must be loaded before sending messages.")

    try:
        encoded = dbc_manager.encode(request.message_name, request.signals)
        message = _build_can_message(encoded)
    except Exception as exc:
        logger.exception("Failed to encode message: %s", exc)
        raise HTTPException(status_code=400, detail=str(exc))

    period_ms = request.period_ms
    try:
        if period_ms:
            period_seconds = period_ms / 1000.0
            task_key = request.task_key or request.message_name
            can_manager.start_periodic(task_key, message, period_seconds)
            _handle_tx_event(request.message_name, encoded, task_key=task_key, period_ms=period_ms)
            return JSONResponse({"status": "periodic", "taskKey": task_key, "periodMs": period_ms})

        can_manager.send(message)
        _handle_tx_event(request.message_name, encoded)
        return JSONResponse({"status": "sent"})

    except CANNotConfiguredError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        logger.exception("Failed to send CAN message: %s", exc)
        raise HTTPException(status_code=500, detail="Internal CAN send error.")


@app.post("/api/messages/stop")
async def stop_message(request: StopTaskRequest) -> Dict[str, Any]:
    can_manager.stop_periodic(request.task_key)
    return {"status": "stopped", "taskKey": request.task_key}


@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket) -> None:
    await broadcaster.connect(websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        await broadcaster.disconnect(websocket)
